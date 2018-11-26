# -*- coding:utf-8 -*-
import logging
import requests
import hashlib
import json
import jsonpath
from openpyxl import Workbook

logging.basicConfig(level=logging.DEBUG,
                    format=u'%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s',
                    datefmt=u'%a, %d %b %Y %H:%M:%S',
                    filename=u'log.log',
                    filemode=u'a+')
console = logging.StreamHandler()
console.setLevel(logging.INFO)

logging.getLogger().addHandler(console)


class HttpClient:
    '''
    Desc:基本的http管理类
    '''

    def __init__(self):
        '''
        Desc:创建http管理类
        Args:
            conf:配置文件
        '''
        self.session = requests.session()
        # self.session.headers['Content-type'] = 'application/x-www-form-urlencoded;charset=UTF-8'
        # requests.packages.urllib3.disable_warnings()

    def httpPost(self, url, params, use_json=False):
        '''
        Desc:发送http请求
        Args:
            url:http请求的url
            params:http请求的内容
        Returns:http请求返回的内容
        '''
        logging.info(u'the post url is %s' % url)
        logging.info(u'the post params is %s' % params)
        logging.info(u'the cookies is %s' % self.session.cookies)
        if use_json:
            response = self.session.post(url, json=params)
        else:
            response = self.session.post(url, data=params)
        content = response.content.decode(u'utf-8')
        logging.debug(u'return content is %s ' % content)
        return content

    def httpGet(self, url):
        '''
        Desc:发送get请求
        Args:
            url:http请求的url
        Returns:http请求返回的内容
        '''
        logging.info(u'the get url is ' + url)
        logging.info(u'the cookies is %s' % self.session.cookies)
        response = self.session.get(url)
        content = response.content.decode(u'utf-8')
        logging.debug(u'the return content is ' + content)
        return content

    def closeSession(self):
        '''
        Desc:关闭session
        '''
        self.session.close()
        self.session = None


class RequireItem:
    def __init__(self, name, path, input_type, range_of_values):
        '''
        :param name:
        :param input_type: text, radio, checkbox
        :param range_of_values:
        '''
        self.name = name
        self.path = path
        self.input_type = input_type
        self.range_of_values = range_of_values
        if input_type not in [u'text', u'radio', u'checkbox']:
            raise Exception(u'%s type error' % name)

    def verify(self, data_json):
        '''
        Desc 验证必填项
        :param data_json:
        :return:
        '''
        logging.info(u'检查[%s]' % self.name)
        actual_data = jsonpath.jsonpath(data_json, self.path)
        if actual_data is None or actual_data is False or len(actual_data) == 0:
            return False
        actual_data = actual_data[0]
        logging.info(u'%s is %s, type is %s, value_range is %s' % (
        self.name, actual_data, self.input_type, self.range_of_values))
        if actual_data is None or len(actual_data.strip()) == 0:
            return False
        actual_data = actual_data.strip()
        self.data = actual_data
        if self.input_type == u'radio':
            values = self.range_of_values.split(u',')
            if actual_data not in values:
                return False
        if self.input_type == u'checkbox':
            values = set(self.range_of_values.split(u','))
            actual_values = set(actual_data.split(u','))
            if not actual_values.issubset(values):
                return False
        return True


class HttpManager:
    def __init__(self, ip, port):
        self.base_url = u'http://%s:%s' % (ip, port)
        self.http_client = HttpClient()

    def login(self, username, password):
        '''
        Desc:登录
        :param username:
        :param password:
        :return:
        '''
        data = {}
        data[u'jOrgNo'] = u''
        data[u'jsu'] = username
        md5 = hashlib.md5()
        md5.update(password)
        data[u'jsp'] = md5.hexdigest()
        login_url = self.base_url + u'/onlineGovQl/j_spring_security_check'
        content = self.http_client.httpPost(login_url, data)
        if u'onlineGovQl/starshine/login.jsp' in content:
            logging.error(u'login %s fail' % username)
            return False
        return True

    def get_user_info(self):
        user_info_url = self.base_url + u'/onlineGovQl/UserController.do?findAdmin=true'
        content = self.http_client.httpPost(user_info_url, {})
        self.user_info_json = json.loads(content)[u'bean']

    def login_out(self):
        login_out_url = self.base_url + u'/onlineGovQl/j_spring_security_logout'
        self.http_client.httpGet(login_out_url)

    def xzqlsx_check(self):
        '''
        Desc:行政权力事项
        行政权利信息：
        权利名称:QL_NAME
        必填项:
        事项编码:DEPT_QL_REG_NO
        权利来源:QL_DEPSTATE:id 单选 1,4,5,6,7,8 当为7 or 8时，委托机关 ENTRUST_NAME必填
        行使层级: USE_LEVEL:id 多选 以,分隔 1,2,3,4,5,6
        行使层级(部门)及内容: USE_LEVEL_C
        行政相对人性质: QL_OBJECT:id 多选 以,分隔 1,2,3
        是否是审核转报事项: IF_AUDIT_TRANSFER: id 单选, 0,1 ,当为1时, 审核转报类型必填 单选 AUDIT_TRANSFER_TYPE:id 1,2,3

        行政权利事项中的业务信息(行政许可)：
        必填项:
        业务名称: YW_NAME
        设定依据: XK_YW_BY_LAW
        行使层级内容: XK_USE_LEVEL_C
        是否委托行使: IF_ENTRUST 单选 "false" "true"
        办件类型: YW_TYPE:text 单选 "承诺件" "即办件"
        法定办结期限: XK_ANTICIPATE_DAY
        承诺办结期限: XK_PROMISE_DAY
        是否收费: CHARGE_FLAG: id 单选 0,1
        法定办结期限单位 XK_ANTICIPATE_TYPE: id 单选 0,1,2,3
        承诺办结期限单位 XK_PROMISE_TYPE： id 单选 0,1,2,3
        受理机构 QL_DEPT_TEXT
        申请条件 CONDITION
        禁止性规定 PROHIBIT_LAW
        数量限制 LIMIT_NUM: id 单选 1,2
        行政相对人性质 YW_QL_OBJECT 多选 1,2,3
        是否涉及投资项目审批 IF_PROJECT_EXAM:id 单选 0,1
        网上办理链接 TRANSACT_URL
        办公地址 TRANSACT_ADDR
        办公时间 TRANSACT_TIME
        涉及中介服务及收费标准 SERVICE_DEPT
        咨询方式 LINK_TEL
        办理进程及结果查询 QUERY_METHOD
        监督投诉渠道 SUPERVISE_TEL
        材料名称 material[] DOCUMENT_NAME
        材料是否必须 material[] IF_NEED 单选 "true" "false"
        常见问题 network_BN_QUESTION_BY_ANSWER[] QUESTION
        常见问题回答 network_BN_QUESTION_BY_ANSWER[] ANSWER
        收费项目 network_BN_CHARGEITEM_INF[] CHARGEITEM_NAME
        收费依据 network_BN_CHARGEITEM_INF[] CHARGEITEM_LAW
        收费标准 network_BN_CHARGEITEM_INF[] CHARGEITEM_STAND

        行政权利事项中的业务信息(行政处罚)：
        必填项:
        业务名称: YW_NAME
        设定依据: XK_YW_BY_LAW
        行使层级内容: XK_USE_LEVEL_C
        是否委托行使: IF_ENTRUST 单选 "false" "true"
        办公地址 TRANSACT_ADDR
        办公时间 TRANSACT_TIME
        处罚结果送达方式 RESULT_SEND_MODE
        行政相对人性质 YW_QL_OBJECT 多选 1,2,3
        咨询方式 LINK_TEL
        监督投诉渠道 SUPERVISE_TEL
        自由裁量标准 违法程度 network_BN_STANDARD_INF[]STA_ILLEGALITY
        自由裁量标准 处罚种类 network_BN_STANDARD_INF[]STA_PUBLISHCLASS
        自由裁量标准 处罚措施 network_BN_STANDARD_INF[]STA_PUBLISHLAW
        常见问题 network_BN_QUESTION_BY_ANSWER[] QUESTION
        常见问题回答 network_BN_QUESTION_BY_ANSWER[] ANSWER

        行政权利事项中的业务信息(行政强制)：
        必填项:
        业务名称: YW_NAME
        设定依据: XK_YW_BY_LAW
        行使层级内容: XK_USE_LEVEL_C
        行政相对人性质 YW_QL_OBJECT 多选 1,2,3
        办公地址 TRANSACT_ADDR
        办公时间 TRANSACT_TIME
        咨询方式 LINK_TEL
        处罚结果送达方式 RESULT_SEND_MODE
        监督投诉渠道 SUPERVISE_TEL
        行政程序 XZ_PROCEDURE
        常见问题 network_BN_QUESTION_BY_ANSWER[] QUESTION
        常见问题回答 network_BN_QUESTION_BY_ANSWER[] ANSWER

        行政权利事项中的业务信息(行政征收)：
        必填项:
        业务名称: YW_NAME
        设定依据: XK_YW_BY_LAW
        行使层级内容: XK_USE_LEVEL_C
        是否委托行使: IF_ENTRUST 单选 "false" "true"
        办件类型: YW_TYPE:text 单选 "承诺件" "即办件"
        是否收费: CHARGE_FLAG: id 单选 0,1
        受理机构 QL_DEPT_TEXT
        申请条件 CONDITION
        禁止性规定 PROHIBIT_LAW
        数量限制 LIMIT_NUM: id 单选 1,2
        行政相对人性质 YW_QL_OBJECT 多选 1,2,3
        是否涉及投资项目审批 IF_PROJECT_EXAM:id 单选 0,1
        网上办理链接 TRANSACT_URL
        办公地址 TRANSACT_ADDR
        办公时间 TRANSACT_TIME
        涉及中介服务及收费标准 SERVICE_DEPT
        咨询方式 LINK_TEL
        办理进程及结果查询 QUERY_METHOD
        监督投诉渠道 SUPERVISE_TEL
        材料名称 material[] DOCUMENT_NAME
        材料是否必须 material[] IF_NEED 单选 "true" "false"
        是否需要电子版 material[] IF_EC_PAGE 单选 "0" "1"
        常见问题 network_BN_QUESTION_BY_ANSWER[] QUESTION
        常见问题回答 network_BN_QUESTION_BY_ANSWER[] ANSWER
        收费项目 network_BN_CHARGEITEM_INF[] CHARGEITEM_NAME
        收费依据 network_BN_CHARGEITEM_INF[] CHARGEITEM_LAW
        收费标准 network_BN_CHARGEITEM_INF[] CHARGEITEM_STAND
        :return:
        '''
        error_list = []
        xzqlsx_map = {u'DEPT_QL_REG_NO': RequireItem(u'事项编码', u'$.DEPT_QL_REG_NO', u'text', u''),
                      u'QL_DEPSTATE': RequireItem(u'权利来源', u'$.QL_DEPSTATE.id', u'radio', u'1,4,5,6,7,8'),
                      u'USE_LEVEL': RequireItem(u'行使层级', u'$.USE_LEVEL.id', u'checkbox', u'1,2,3,4,5,6'),
                      u'USE_LEVEL_C': RequireItem(u'行使层级(部门)及内容', u'$.USE_LEVEL_C', u'text', u''),
                      u'QL_OBJECT': RequireItem(u'行政相对人性质', u'$.QL_OBJECT.id', u'checkbox', u'1,2,3'),
                      u'IF_AUDIT_TRANSFER': RequireItem(u'是否是审核转报事项', u'$.IF_AUDIT_TRANSFER.id', u'radio', u'0,1')}
        xzqlsx_url = self.base_url + u'/onlineGovQl/QlywConController.do?findQlywMain=true'
        params = {u'pageStart': u'1', u'limit': u'-1', u'QL_DEPT[id]': self.user_info_json[u'idOrg'],
                  u'QL_DEPT[text]': self.user_info_json[u'orgName'],
                  u'QL_REG_NO': u'', u'QL_CODE': u'', u'QL_NAME': u'', u'QL_SHORT_NAME': u'',
                  u'IF_AUDIT_TRANSFER[id]': u'all', u'IF_AUDIT_TRANSFER[text]': u'全部', u'NUM': u'',
                  u'B[YW_OPERATE][id]': u'', u'B[YW_OPERATE][text]': u'', u'B[DEPT_YW_NUM]': u'', u'B[YE_NAME]': u'',
                  u'B[TRANSACT_DEP]': u'', u'B[YE_FILENUM]': u''}
        content = self.http_client.httpPost(xzqlsx_url, params)
        result_json = json.loads(content)
        logging.info(u'total count is %s' % result_json[u'total'])
        logging.info(u'total items is %s' % len(result_json[u'items']))
        for item in result_json[u'items']:
            logging.info(u'权利名称 is %s' % item[u'QL_NAME'])
            if item[u'isParent'] != u'true':
                continue
            errors = []
            for key, require_item in xzqlsx_map.iteritems():
                if not require_item.verify(item):
                    errors.append(u'%s 不正确' % require_item.name)
            if xzqlsx_map[u'QL_DEPSTATE'].data in [u'7', u'8']:
                require_item = RequireItem(u'委托机关 ', u'$.ENTRUST_NAME', u'text', u'')
                if not require_item.verify(item):
                    errors.append(u'%s 不正确' % require_item.name)
            if xzqlsx_map[u'IF_AUDIT_TRANSFER'].data in [u'1']:
                require_item = RequireItem(u'审核转报类型 ', u'$.AUDIT_TRANSFER_TYPE.id', u'radio', u'1,2,3')
                if not require_item.verify(item):
                    errors.append(u'[%s]不正确' % require_item.name)

            params = {u'idt': item[u'IDDEPT_QL_INF'], u'sEcho': u'1', u'pageStart': u'1', u'limit': u'-1',
                      u'iSortCol': u'', u'sSortDir': u''}
            ywxx_url = self.base_url + u'/onlineGovQl/QlywConController.do?findQlyw=true&start=aoData["iDisplayStart"]'
            content = self.http_client.httpPost(ywxx_url, params)
            ywxxs_json = json.loads(content)
            if item[u'QL_KIND'] == u'1':
                ywxx_map = {u'YW_NAME': RequireItem(u'业务名称', u'$.YW_NAME', u'text', u''),
                            u'XK_YW_BY_LAW': RequireItem(u'设定依据', u'$.XK_YW_BY_LAW', u'text', u''),
                            u'XK_USE_LEVEL_C': RequireItem(u'行使层级内容', u'$.XK_USE_LEVEL_C', u'text', u''),
                            u'IF_ENTRUST': RequireItem(u'是否委托行使', u'$.IF_ENTRUST', u'radio', u'true,false'),
                            u'YW_TYPE': RequireItem(u'办件类型', u'$.YW_TYPE.text', u'radio', u'承诺件,即办件'),
                            u'XK_ANTICIPATE_DAY': RequireItem(u'法定办结期限', u'$.XK_ANTICIPATE_DAY', u'text', u''),
                            u'XK_PROMISE_DAY': RequireItem(u'承诺办结期限', u'$.XK_PROMISE_DAY', u'text', u''),
                            u'CHARGE_FLAG': RequireItem(u'是否收费', u'$.CHARGE_FLAG.id', u'radio', u'0,1'),
                            u'XK_ANTICIPATE_TYPE': RequireItem(u'法定办结期限单位', u'$.XK_ANTICIPATE_TYPE.id', u'radio',
                                                               u'0,1,2,3'),
                            u'XK_PROMISE_TYPE': RequireItem(u'承诺办结期限单位', u'$.XK_PROMISE_TYPE.id', u'radio', u'0,1,2,3'),
                            u'QL_DEPT_TEXT': RequireItem(u'受理机构', u'$.QL_DEPT_TEXT', u'text', u''),
                            u'CONDITION': RequireItem(u'申请条件', u'$.CONDITION', u'text', u''),
                            u'PROHIBIT_LAW': RequireItem(u'禁止性规定', u'$.PROHIBIT_LAW', u'text', u''),
                            u'LIMIT_NUM': RequireItem(u'数量限制', u'$.LIMIT_NUM.id', u'radio', u'1,2'),
                            u'YW_QL_OBJECT': RequireItem(u'行政相对人性质', u'$.YW_QL_OBJECT', u'checkbox', u'1,2,3'),
                            u'IF_PROJECT_EXAM': RequireItem(u'是否涉及投资项目审批', u'$.IF_PROJECT_EXAM.id', u'radio', u'0,1'),
                            u'TRANSACT_URL': RequireItem(u'网上办理链接', u'$.TRANSACT_URL', u'text', u''),
                            u'TRANSACT_ADDR': RequireItem(u'办公地址', u'$.TRANSACT_ADDR', u'text', u''),
                            u'TRANSACT_TIME': RequireItem(u'办公时间', u'$.TRANSACT_TIME', u'text', u''),
                            u'SERVICE_DEPT': RequireItem(u'涉及中介服务及收费标准', u'$.SERVICE_DEPT', u'text', u''),
                            u'LINK_TEL': RequireItem(u'咨询方式', u'$.LINK_TEL', u'text', u''),
                            u'QUERY_METHOD': RequireItem(u'办理进程及结果查询', u'$.QUERY_METHOD', u'text', u''),
                            u'SUPERVISE_TEL': RequireItem(u'监督投诉渠道', u'$.SUPERVISE_TEL', u'text', u'')
                            }
                ywxx_material_map = {u'DOCUMENT_NAME': RequireItem(u'材料名称', u'$.DOCUMENT_NAME', u'text', u''),
                                     u'IF_NEED': RequireItem(u'材料是否必须', u'$.IF_NEED', u'radio', u'true,false'),
                                     u'IF_EC_PAGE': RequireItem(u'是否需要电子版', u'$.IF_EC_PAGE', u'radio', u'0,1')}
                ywxx_question_map = {u'QUESTION': RequireItem(u'常见问题', u'$.QUESTION', u'text', u''),
                                     u'ANSWER': RequireItem(u'常见问题回答', u'$.ANSWER', u'text', u'')}
                ywxx_charge_map = {u'CHARGEITEM_NAME': RequireItem(u'收费项目', u'$.CHARGEITEM_NAME', u'text', u''),
                                   u'CHARGEITEM_LAW': RequireItem(u'收费依据', u'$.CHARGEITEM_LAW', u'text', u''),
                                   u'CHARGEITEM_STAND': RequireItem(u'收费标准', u'$.CHARGEITEM_STAND', u'text', u'')}
                ywxx_standard_map = {}
            elif item[u'QL_KIND'] == u'2':
                ywxx_map = {u'YW_NAME': RequireItem(u'业务名称', u'$.YW_NAME', u'text', u''),
                            u'XK_YW_BY_LAW': RequireItem(u'设定依据', u'$.XK_YW_BY_LAW', u'text', u''),
                            u'XK_USE_LEVEL_C': RequireItem(u'行使层级内容', u'$.XK_USE_LEVEL_C', u'text', u''),
                            u'IF_ENTRUST': RequireItem(u'是否委托行使', u'$.IF_ENTRUST', u'radio', u'true,false'),
                            u'YW_QL_OBJECT': RequireItem(u'行政相对人性质', u'$.YW_QL_OBJECT', u'checkbox', u'1,2,3'),
                            u'TRANSACT_ADDR': RequireItem(u'办公地址', u'$.TRANSACT_ADDR', u'text', u''),
                            u'TRANSACT_TIME': RequireItem(u'办公时间', u'$.TRANSACT_TIME', u'text', u''),
                            u'LINK_TEL': RequireItem(u'咨询方式', u'$.LINK_TEL', u'text', u''),
                            u'RESULT_SEND_MODE': RequireItem(u'处罚结果送达方式', u'$.RESULT_SEND_MODE', u'text', u''),
                            u'SUPERVISE_TEL': RequireItem(u'监督投诉渠道', u'$.SUPERVISE_TEL', u'text', u'')
                            }
                ywxx_material_map = {}
                ywxx_question_map = {u'QUESTION': RequireItem(u'常见问题', u'$.QUESTION', u'text', u''),
                                     u'ANSWER': RequireItem(u'常见问题回答', u'$.ANSWER', u'text', u'')}
                ywxx_charge_map = {}
                ywxx_standard_map = {u'STA_ILLEGALITY': RequireItem(u'违法程度', u'$.STA_ILLEGALITY', u'text', u''),
                                     u'STA_PUBLISHCLASS': RequireItem(u'处罚种类', u'$.STA_PUBLISHCLASS', u'text', u''),
                                     u'STA_PUBLISHLAW': RequireItem(u'处罚措施', u'$.STA_PUBLISHLAW', u'text', u'')}
            elif item[u'QL_KIND'] in [u'3', u'9']:
                ywxx_map = {u'YW_NAME': RequireItem(u'业务名称', u'$.YW_NAME', u'text', u''),
                            u'XK_YW_BY_LAW': RequireItem(u'设定依据', u'$.XK_YW_BY_LAW', u'text', u''),
                            u'XK_USE_LEVEL_C': RequireItem(u'行使层级内容', u'$.XK_USE_LEVEL_C', u'text', u''),
                            u'YW_QL_OBJECT': RequireItem(u'行政相对人性质', u'$.YW_QL_OBJECT', u'checkbox', u'1,2,3'),
                            u'TRANSACT_ADDR': RequireItem(u'办公地址', u'$.TRANSACT_ADDR', u'text', u''),
                            u'TRANSACT_TIME': RequireItem(u'办公时间', u'$.TRANSACT_TIME', u'text', u''),
                            u'XZ_PROCEDURE': RequireItem(u'行政程序', u'$.XZ_PROCEDURE', u'text', u''),
                            u'LINK_TEL': RequireItem(u'咨询方式', u'$.LINK_TEL', u'text', u''),
                            u'RESULT_SEND_MODE': RequireItem(u'结果送达方式', u'$.RESULT_SEND_MODE', u'text', u''),
                            u'SUPERVISE_TEL': RequireItem(u'监督投诉渠道', u'$.SUPERVISE_TEL', u'text', u'')
                            }
                ywxx_material_map = {}
                ywxx_question_map = {u'QUESTION': RequireItem(u'常见问题', u'$.QUESTION', u'text', u''),
                                     u'ANSWER': RequireItem(u'常见问题回答', u'$.ANSWER', u'text', u'')}
                ywxx_charge_map = {}
                ywxx_standard_map = {}
            elif item[u'QL_KIND'] in [u'4', u'5', u'6', u'7', u'8', u'10']:
                ywxx_map = {u'YW_NAME': RequireItem(u'业务名称', u'$.YW_NAME', u'text', u''),
                            u'XK_YW_BY_LAW': RequireItem(u'设定依据', u'$.XK_YW_BY_LAW', u'text', u''),
                            u'XK_USE_LEVEL_C': RequireItem(u'行使层级内容', u'$.XK_USE_LEVEL_C', u'text', u''),
                            u'IF_ENTRUST': RequireItem(u'是否委托行使', u'$.IF_ENTRUST', u'radio', u'true,false'),
                            u'YW_TYPE': RequireItem(u'办件类型', u'$.YW_TYPE.text', u'radio', u'承诺件,即办件'),
                            u'CHARGE_FLAG': RequireItem(u'是否收费', u'$.CHARGE_FLAG.id', u'radio', u'0,1'),
                            u'QL_DEPT_TEXT': RequireItem(u'受理机构', u'$.QL_DEPT_TEXT', u'text', u''),
                            u'CONDITION': RequireItem(u'申请条件', u'$.CONDITION', u'text', u''),
                            u'PROHIBIT_LAW': RequireItem(u'禁止性规定', u'$.PROHIBIT_LAW', u'text', u''),
                            u'LIMIT_NUM': RequireItem(u'数量限制', u'$.LIMIT_NUM.id', u'radio', u'1,2'),
                            u'YW_QL_OBJECT': RequireItem(u'行政相对人性质', u'$.YW_QL_OBJECT', u'checkbox', u'1,2,3'),
                            u'IF_PROJECT_EXAM': RequireItem(u'是否涉及投资项目审批', u'$.IF_PROJECT_EXAM.id', u'radio', u'0,1'),
                            u'TRANSACT_URL': RequireItem(u'网上办理链接', u'$.TRANSACT_URL', u'text', u''),
                            u'TRANSACT_ADDR': RequireItem(u'办公地址', u'$.TRANSACT_ADDR', u'text', u''),
                            u'TRANSACT_TIME': RequireItem(u'办公时间', u'$.TRANSACT_TIME', u'text', u''),
                            u'SERVICE_DEPT': RequireItem(u'涉及中介服务及收费标准', u'$.SERVICE_DEPT', u'text', u''),
                            u'LINK_TEL': RequireItem(u'咨询方式', u'$.LINK_TEL', u'text', u''),
                            u'QUERY_METHOD': RequireItem(u'办理进程及结果查询', u'$.QUERY_METHOD', u'text', u''),
                            u'SUPERVISE_TEL': RequireItem(u'监督投诉渠道', u'$.SUPERVISE_TEL', u'text', u'')
                            }
                ywxx_material_map = {u'DOCUMENT_NAME': RequireItem(u'材料名称', u'$.DOCUMENT_NAME', u'text', u''),
                                     u'IF_NEED': RequireItem(u'材料是否必须', u'$.IF_NEED', u'radio', u'true,false'),
                                     u'IF_EC_PAGE': RequireItem(u'是否需要电子版', u'$.IF_EC_PAGE', u'radio', u'0,1')}
                ywxx_question_map = {u'QUESTION': RequireItem(u'常见问题', u'$.QUESTION', u'text', u''),
                                     u'ANSWER': RequireItem(u'常见问题回答', u'$.ANSWER', u'text', u'')}
                ywxx_charge_map = {u'CHARGEITEM_NAME': RequireItem(u'收费项目', u'$.CHARGEITEM_NAME', u'text', u''),
                                   u'CHARGEITEM_LAW': RequireItem(u'收费依据', u'$.CHARGEITEM_LAW', u'text', u''),
                                   u'CHARGEITEM_STAND': RequireItem(u'收费标准', u'$.CHARGEITEM_STAND', u'text', u'')}
                ywxx_standard_map = {}

            for ywxx in ywxxs_json[u'items']:
                params = {u'idt': ywxx[u'IDDEPT_YW_INF'], u'idtlog': ywxx[u'IDDEPT_YW_INFOLD'], u'showTable': u'XK'}
                ywxx_detail_url = self.base_url + u'/onlineGovQl/QlywConController.do?findServiseData=true'
                content = self.http_client.httpPost(ywxx_detail_url, params)
                ywxx_detail_json = json.loads(content)[u'bean'][u'data'][0]
                logging.info(u'检查业务信息[%s]' % ywxx_detail_json[u'YW_NAME'])
                for key, require_item in ywxx_map.iteritems():
                    if not require_item.verify(ywxx_detail_json):
                        errors.append(u'业务名称[%s]的[%s]不正确' % (ywxx_detail_json[u'YW_NAME'], require_item.name))
                for material_item in ywxx_detail_json[u'material']:
                    for key, require_item in ywxx_material_map.iteritems():
                        if not require_item.verify(material_item):
                            errors.append(u'第[%s]条材料的[%s]不正确' % (material_item[u'ORD'], require_item.name))
                for question_item in ywxx_detail_json[u'network_BN_QUESTION_BY_ANSWER']:
                    for key, require_item in ywxx_question_map.iteritems():
                        if not require_item.verify(question_item):
                            errors.append(u'第[%s]条常见问题的[%s]不正确' % (question_item[u'ORD'], require_item.name))
                for charge_item in ywxx_detail_json[u'network_BN_CHARGEITEM_INF']:
                    for key, require_item in ywxx_charge_map.iteritems():
                        if not require_item.verify(charge_item):
                            errors.append(u'第[%s]条收费项目的[%s]不正确' % (charge_item[u'ORD'], require_item.name))
                for standard_item in ywxx_detail_json[u'network_BN_STANDARD_INF']:
                    for key, require_item in ywxx_standard_map.iteritems():
                        if not require_item.verify(standard_item):
                            errors.append(u'第[%s]条自由裁量标准的[%s]不正确' % (standard_item[u'ORD'], require_item.name))
            if errors:
                error_item = {u'name': item[u'QL_NAME'], u'id': item[u'DEPT_QL_REG_NO'], u'errors': errors}
                error_list.append(error_item)
        logging.info(error_list)
        return error_list

    def kzxx_check(self):
        '''
        Desc:扩展信息检查
        进驻大厅 IF_JZ_HALL 单选 0,1
        网上办理深度 ONLINE_BL_SD 单选 1,2,3,4
        网上缴费 ONLINE_PAY 单选 0,1
        网上申报 IF_ONLINE_SB 单选 0,1
        网上全程办结 ONLINE_QC_BJ 单选 0,1
        快递服务 ONLINE_EMS 单选 0,1
        在线申请建设 ZXSP_TYPE 1,2
        :return:
        '''
        error_list = []
        kzxx_map = {u'IF_JZ_HALL': RequireItem(u'进驻大厅', u'$.IF_JZ_HALL', u'radio', u'0,1'),
                    u'ONLINE_BL_SD': RequireItem(u'网上办理深度', u'$.ONLINE_BL_SD', u'radio', u'1,2,3,4'),
                    u'ONLINE_PAY': RequireItem(u'网上缴费', u'$.ONLINE_PAY', u'radio', u'0,1'),
                    u'IF_ONLINE_SB': RequireItem(u'网上申报', u'$.IF_ONLINE_SB', u'radio', u'0,1'),
                    u'ONLINE_QC_BJ': RequireItem(u'网上全程办结', u'$.ONLINE_QC_BJ', u'radio', u'0,1'),
                    u'ONLINE_EMS': RequireItem(u'是否是审核转报事项', u'$.ONLINE_EMS', u'radio', u'0,1'),
                    u'ZXSP_TYPE': RequireItem(u'在线申请建设', u'$.ZXSP_TYPE', u'radio', u'1,2')}
        kzxx_url = self.base_url + u'/onlineGovQl/OnlineServiceController.do?findMainTables=true&start=aoData["iDisplayStart"]'
        params = {u'XZQH[id]': self.user_info_json[u'areaAdress'], u'XZQH[text]': self.user_info_json[u'areaName'],
                  u'BMMC[id]': self.user_info_json[u'idOrg'], u'BMMC[text]': self.user_info_json[u'orgName'],
                  u'QL_KIND[id]': u'',
                  u'QL_KIND[text]': u'', u'QL_REG_NO': u'', u'QL_NAME': u'', u'IF_AUDIT_TRANSFER[id]': u'all',
                  u'IF_AUDIT_TRANSFER[text]': u'全部', u'AUDIT_TRANSFER_TYPE[id]': u'', u'AUDIT_TRANSFER_TYPE[text]': u'',
                  u'YW_NAME': u'', u'AAA': u'qweqwe', u'TRANSACT_URL': u'-1', u'STATE[id]': u'', u'STATE[text]': u'',
                  u'sEcho': u'1',
                  u'pageStart': u'1', u'limit': u'-1', u'iSortCol': u'', u'sSortDir': u''}
        content = self.http_client.httpPost(kzxx_url, params)
        result_json = json.loads(content)
        logging.info(u'total count is %s' % result_json[u'total'])
        logging.info(u'total items is %s' % len(result_json[u'items']))
        for item in result_json[u'items']:
            errors = []
            logging.info(u'检查权利名称[%s]的业务名称[%s]' % (item[u'QL_NAME'], item[u'YW_NAME']))
            for key, require_item in kzxx_map.iteritems():
                if not require_item.verify(item):
                    errors.append(u'%s 不正确' % require_item.name)
            if errors:
                error_item = {u'name': u'权利名称[%s]的业务名称[%s]' % (item[u'QL_NAME'], item[u'YW_NAME']),
                              u'id': item[u'IDDEPT_YW_INF'], u'errors': errors}
                error_list.append(error_item)
        logging.info(error_list)
        return error_list

    def ggfw_check(self):
        '''
        Desc:公共服务检查
        服务编码: DEPT_QL_REG_NO
        服务名称: QL_NAME
        必填项:
        服务领域: G_FW_RANGE 单选 1-35, 99
        服务方式: G_FW_MODE 单选 1,2
        服务内容: G_FW_CONTENT
        服务依据: QL_BY_LAW

        业务信息:
        服务方式为主动服务:
        必填项:
        业务名称 YW_NAME
        承办机构 QL_DEPT_TEXT
        经办机构（处室） TRANSACT_DEP
        数量限制 LIMIT_NUM.id 单选 1,2
        服务对象 YW_QL_OBJECT 多选 1,2,3
        办理地点 TRANSACT_ADDR
        工作时间 TRANSACT_TIME
        咨询方式 LINK_TEL
        监督投诉方式 SUPERVISE_TEL

        服务方式为依申请服务:
        必填项:
        业务名称 YW_NAME
        办理依据 XK_YW_BY_LAW
        办件类型 YW_TYPE: text 单选 "承诺件" "即办件"
        是否收费 CHARGE_FLAG: id 单选 0,1
        承办机构 QL_DEPT_TEXT
        经办机构（处室）TRANSACT_DEP
        数量限制 LIMIT_NUM.id 单选 1,2
        服务对象 YW_QL_OBJECT 多选 1,2,3
        在线办理网址 TRANSACT_URL
        办理地点 TRANSACT_ADDR
        办理时间 TRANSACT_TIME
        咨询方式 LINK_TEL
        监督投诉方式 SUPERVISE_TEL
        材料名称 material[] DOCUMENT_NAME
        材料是否必须 material[] IF_NEED 单选 "true" "false"
        是否需要电子版 material[] IF_EC_PAGE 单选 "0" "1"
        常见问题 network_BN_QUESTION_BY_ANSWER[] QUESTION
        常见问题回答 network_BN_QUESTION_BY_ANSWER[] ANSWER
        收费项目 network_BN_CHARGEITEM_INF[] CHARGEITEM_NAME
        收费依据 network_BN_CHARGEITEM_INF[] CHARGEITEM_LAW
        收费标准 network_BN_CHARGEITEM_INF[] CHARGEITEM_STAND
        扩展信息 到现场次数 pubservice_yw_expand:DAO_XC_NUM 单选 0,1,2,99 当不为0时，见面法律依据 必填 DAO_XC_LAW
        :return:
        '''
        error_list = []
        ggfw_url = self.base_url + u'/onlineGovQl/PubServiceYwController.do?findQlywMain=true'
        params = {u'pageStart': u'1', u'limit': u'-1',
                  u'QL_DEPT[id]': self.user_info_json[u'idOrg'], u'QL_DEPT[text]': self.user_info_json[u'orgName'],
                  u'QL_REG_NO': u'',
                  u'QL_CODE': u'', u'QL_NAME': u'', u'QL_SHORT_NAME': u'', u'IF_AUDIT_TRANSFER[id]': u'all',
                  u'IF_AUDIT_TRANSFER[text]': u'全部', u'NUM': u'', u'B[YW_OPERATE][id]': u'',
                  u'B[YW_OPERATE][text]': u'', u'B[DEPT_YW_NUM]': u'', u'B[YE_NAME]': u'', u'B[TRANSACT_DEP]': u'',
                  u'B[YE_FILENUM]': u'', u'G_FW_RANGE[id]': u'',
                  u'G_FW_RANGE[text]': u'',
                  u'G_FW_MODE[id]': u'', u'G_FW_MODE[text]': u''}

        content = self.http_client.httpPost(ggfw_url, params)
        result_json = json.loads(content)
        array_str = map(unicode, range(1, 36))
        array_str.append(u'99')
        fw_range = u','.join(array_str)
        ggfw_map = {u'G_FW_RANGE': RequireItem(u'服务领域', u'$.G_FW_RANGE', u'radio', fw_range),
                    u'G_FW_MODE': RequireItem(u'服务方式', u'$.G_FW_MODE', u'radio', u'1,2'),
                    u'G_FW_CONTENT': RequireItem(u'服务内容', u'$.G_FW_CONTENT', u'text', u''),
                    u'QL_BY_LAW': RequireItem(u'服务依据', u'$.QL_BY_LAW', u'text', u'')}
        for ggfw_item in result_json[u'items']:
            errors = []
            ggfw_detail_url = self.base_url + u'/onlineGovQl/PubServiceYwController.do?findQlMain=true'
            params = {u'idt': ggfw_item[u'IDDEPT_QL_INF']}
            content = self.http_client.httpPost(ggfw_detail_url, params)
            ggfw_detail_json = json.loads(content)[u'items'][0]
            logging.info(u'检查服务名称[%s]' % ggfw_detail_json[u'QL_NAME'])
            for key, require_item in ggfw_map.iteritems():
                if not require_item.verify(ggfw_detail_json):
                    errors.append(u'%s 不正确' % require_item.name)
            ywxx_url = self.base_url + u'/onlineGovQl/PubServiceYwController.do?findQlyw=true&start=aoData["iDisplayStart"]'
            params = {u'idt': ggfw_item[u'IDDEPT_QL_INF'], u'sEcho': u'1', u'pageStart': u'1', u'limit': u'-1',
                      u'iSortCol': u'', u'sSortDir': u''}
            content = self.http_client.httpPost(ywxx_url, params)
            ywxxs_json = json.loads(content)[u'items']
            if ggfw_detail_json[u'G_FW_MODE'] == u'2':
                ywxx_map = {u'YW_NAME': RequireItem(u'业务名称', u'$.YW_NAME', u'text', u''),
                            u'QL_DEPT_TEXT': RequireItem(u'承办机构', u'$.QL_DEPT_TEXT', u'text', u''),
                            u'TRANSACT_DEP': RequireItem(u'经办机构（处室）', u'$.TRANSACT_DEP', u'text', u''),
                            u'LIMIT_NUM': RequireItem(u'数量限制', u'$.LIMIT_NUM.id', u'radio', u'1,2'),
                            u'YW_QL_OBJECT': RequireItem(u'服务对象', u'$.YW_QL_OBJECT', u'checkbox', u'1,2,3'),
                            u'TRANSACT_ADDR': RequireItem(u'办理地点', u'$.TRANSACT_ADDR', u'text', u''),
                            u'TRANSACT_TIME': RequireItem(u'工作时间', u'$.TRANSACT_TIME', u'text', u''),
                            u'LINK_TEL': RequireItem(u'咨询方式', u'$.LINK_TEL', u'text', u''),
                            u'SUPERVISE_TEL': RequireItem(u'监督投诉方式', u'$.SUPERVISE_TEL', u'text', u'')}
                ywxx_material_map = {}
                ywxx_question_map = {}
                ywxx_charge_map = {}
                ywxx_expand_map = {}
            else:
                ywxx_map = {u'YW_NAME': RequireItem(u'业务名称', u'$.YW_NAME', u'text', u''),
                            u'XK_YW_BY_LAW': RequireItem(u'办理依据', u'$.XK_YW_BY_LAW', u'text', u''),
                            u'YW_TYPE': RequireItem(u'办件类型', u'$.YW_TYPE.text', u'radio', u'承诺件,即办件'),
                            u'CHARGE_FLAG': RequireItem(u'是否收费', u'$.CHARGE_FLAG.id', u'radio', u'0,1'),
                            u'QL_DEPT_TEXT': RequireItem(u'承办机构', u'$.QL_DEPT_TEXT', u'text', u''),
                            u'TRANSACT_DEP': RequireItem(u'经办机构（处室）', u'$.TRANSACT_DEP', u'text', u''),
                            u'LIMIT_NUM': RequireItem(u'数量限制', u'$.LIMIT_NUM.id', u'radio', u'1,2'),
                            u'YW_QL_OBJECT': RequireItem(u'服务对象', u'$.YW_QL_OBJECT', u'checkbox', u'1,2,3'),
                            u'TRANSACT_URL': RequireItem(u'在线办理网址', u'$.TRANSACT_URL', u'text', u''),
                            u'TRANSACT_ADDR': RequireItem(u'办理地点', u'$.TRANSACT_ADDR', u'text', u''),
                            u'TRANSACT_TIME': RequireItem(u'工作时间', u'$.TRANSACT_TIME', u'text', u''),
                            u'LINK_TEL': RequireItem(u'咨询方式', u'$.LINK_TEL', u'text', u''),
                            u'SUPERVISE_TEL': RequireItem(u'监督投诉方式', u'$.SUPERVISE_TEL', u'text', u'')}
                ywxx_material_map = {u'DOCUMENT_NAME': RequireItem(u'材料名称', u'$.DOCUMENT_NAME', u'text', u''),
                                     u'IF_NEED': RequireItem(u'材料是否必须', u'$.IF_NEED', u'radio', u'true,false'),
                                     u'IF_EC_PAGE': RequireItem(u'是否需要电子版', u'$.IF_EC_PAGE', u'radio', u'0,1')}
                ywxx_question_map = {u'QUESTION': RequireItem(u'常见问题', u'$.QUESTION', u'text', u''),
                                     u'ANSWER': RequireItem(u'常见问题回答', u'$.ANSWER', u'text', u'')}
                ywxx_charge_map = {u'CHARGEITEM_NAME': RequireItem(u'收费项目', u'$.CHARGEITEM_NAME', u'text', u''),
                                   u'CHARGEITEM_LAW': RequireItem(u'收费依据', u'$.CHARGEITEM_LAW', u'text', u''),
                                   u'CHARGEITEM_STAND': RequireItem(u'收费标准', u'$.CHARGEITEM_STAND', u'text', u'')}
                ywxx_expand_map = {u'DAO_XC_NUM': RequireItem(u'到现场次数', u'$.pubservice_yw_expand.DAO_XC_NUM', u'radio',
                                                              u'0,1,2,99')}
            for ywxx in ywxxs_json:
                ywxx_detail_url = self.base_url + u'/onlineGovQl/PubServiceYwController.do?findServiseData=true'
                params = {u'idt': ywxx[u'IDDEPT_YW_INF'], u'idtlog': ywxx[u'IDDEPT_YW_INFOLD'], u'showTable': u'XK'}
                content = self.http_client.httpPost(ywxx_detail_url, params)
                ywxx_detail_json = json.loads(content)[u'bean'][u'data'][0]
                logging.info(u'检查业务信息[%s]' % ywxx_detail_json[u'YW_NAME'])
                for key, require_item in ywxx_map.iteritems():
                    if not require_item.verify(ywxx_detail_json):
                        errors.append(u'业务名称[%s]的[%s]不正确' % (ywxx_detail_json[u'YW_NAME'], require_item.name))
                for material_item in ywxx_detail_json[u'material']:
                    for key, require_item in ywxx_material_map.iteritems():
                        if not require_item.verify(material_item):
                            errors.append(u'第[%s]条材料的[%s]不正确' % (material_item[u'ORD'], require_item.name))
                for question_item in ywxx_detail_json[u'network_BN_QUESTION_BY_ANSWER']:
                    for key, require_item in ywxx_question_map.iteritems():
                        if not require_item.verify(question_item):
                            errors.append(u'第[%s]条常见问题的[%s]不正确' % (question_item[u'ORD'], require_item.name))
                for charge_item in ywxx_detail_json[u'network_BN_CHARGEITEM_INF']:
                    for key, require_item in ywxx_charge_map.iteritems():
                        if not require_item.verify(charge_item):
                            errors.append(u'第[%s]条收费项目的[%s]不正确' % (charge_item[u'ORD'], require_item.name))
                for key, require_item in ywxx_expand_map.iteritems():
                    if not require_item.verify(ywxx_detail_json):
                        errors.append(u'扩展信息的[%s]不正确' % require_item.name)
                    if ywxx_detail_json[u'pubservice_yw_expand'][u'DAO_XC_NUM'] != u'0':
                        special_require_item = RequireItem(u'见面法律依据', u'$.pubservice_yw_expand.DAO_XC_LAW', u'text',
                                                           u'')
                        if not special_require_item.verify(ywxx_detail_json):
                            errors.append(u'扩展信息的[%s]不正确' % special_require_item.name)
            if errors:
                error_item = {u'name': ggfw_item[u'QL_NAME'], u'id': ggfw_item[u'DEPT_QL_REG_NO'], u'errors': errors}
                error_list.append(error_item)
        logging.info(error_list)
        return error_list

    def check(self, username, password):
        logging.info(u'start to check username %s' % username)
        if not self.login(username, password):
            return
        self.get_user_info()
        xzql_error_list = self.xzqlsx_check()
        kzxx_error_list = self.kzxx_check()
        ggfw_error_list = self.ggfw_check()
        self.login_out()
        if xzql_error_list or kzxx_error_list:
            wb = Workbook()
            if xzql_error_list:
                ws = wb.create_sheet(u'行政权利事项')
                ws.append([u'权利名称', u'权利id', u'错误信息'])
                for error in xzql_error_list:
                    error_message = u'\r\n'.join(error[u'errors'])
                    ws.append([error[u'name'], error[u'id'], error_message])
            if kzxx_error_list:
                ws = wb.create_sheet(u'拓展信息')
                ws.append([u'名称', u'id', u'错误信息'])
                for error in kzxx_error_list:
                    error_message = u'\r\n'.join(error[u'errors'])
                    ws.append([error[u'name'], error[u'id'], error_message])
            if ggfw_error_list:
                ws = wb.create_sheet(u'公共服务')
                ws.append([u'服务名称', u'服务编码', u'错误信息'])
                for error in ggfw_error_list:
                    error_message = u'\r\n'.join(error[u'errors'])
                    ws.append([error[u'name'], error[u'id'], error_message])
            wb.save(u'%s.xlsx' % username)


if __name__ == u'__main__':
    http_manager = HttpManager(u'221.226.253.51', u'5065')
    with open(u'username.txt') as fd:
        lines = fd.readlines()
    for line in lines:
        line = line.strip()
        username, password = line.split(u',')
        http_manager.check(username, password)


